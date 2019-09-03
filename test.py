#coding: utf-8
import pymysql
import sys
import uuid
import requests
import hashlib
import time
import json
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
from openpyxl.styles import PatternFill, colors


def handle(cell):
    if cell is None or cell.value is None:
        return ""
    return str(cell.value)


if __name__ == "__main__":

    path_src = "Language.xlsx"
    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(
        filter(lambda x: x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    start_row = 3

    for row in range(start_row, sheet_src.max_row + 1):
        lanId = int(sheet_src["A{}".format(row)].value)
        cn = handle(sheet_src["B{}".format(row)])
        en = handle(sheet_src["C{}".format(row)])
        zh = handle(sheet_src["D{}".format(row)])
        jp = handle(sheet_src["E{}".format(row)])
        ko = handle(sheet_src["F{}".format(row)])


# #连接数据库
# db = pymysql.connect("localhost","root","","language")

# #使用cursor()方法创建一个游标对象
# cursor = db.cursor()

# #使用execute()方法执行SQL语句
# cursor.execute("SELECT * FROM language")

# #使用fetall()获取全部数据
# data = cursor.fetchall()

# #打印获取到的数据
# print(data)

# #关闭游标和数据库的连接
# cursor.close()
# db.close()
