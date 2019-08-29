from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
import os
import Language_pb2
from Language_pb2 import LanguageTable, LanguageInfo
import shutil


class Table:
    def __init__(self):
        self.error_txt = None
        self.table_dic = {}  # cn -> id
        self.max_id_value = 0
        self.max_row = 0
        self.id_start_row = 3
        self.backup_order = 0
        self.max_backup_count = 30
        self.backup_folder = "backup"

    def load(self, path, shert_name):
        self.path_src = path

        try:
            self.workbook_src = load_workbook(self.path_src, data_only=True)
            self.sheet_src = self.workbook_src[shert_name]
            self.max_row = self.sheet_src.max_row
            for i in range(self.id_start_row, self.max_row + 1):
                cn_value = self.get_lan(i,  Language_pb2.ch)
                id_value = self.get_id(i)
                self.max_id_value = max(self.max_id_value, id_value)
                if cn_value in self.table_dic:
                    self.error_txt = "简体中文 \"{}\" 存在重复！".format(cn_value)
                    break
                self.table_dic[cn_value] = id_value
        except Exception as e:
            self.error_txt = "解析excel表时报错! {}".format(e)


    def insert(self, cn):
        newid = self.max_id_value + 1

        self.table_dic[cn] = newid
        self.max_id_value += 1
        self.max_row += 1

        self.sheet_src["{}{}".format('B', self.max_row)].value = cn
        self.sheet_src["{}{}".format('A', self.max_row)].value = newid

        return newid

    def remove(self, cn):
        if cn not in self.table_dic:
            return

        lanId = self.table_dic[cn]

        del self.table_dic[cn]

        for i in range(self.id_start_row, self.max_row + 1):
            id_value = self.get_id(i)
            if id_value == lanId:
                self.sheet_src.delete_rows(i, 1)
                break

        self.max_row = self.sheet_src.max_row

    def change(self, lanId, oldcn, cn):
        if oldcn not in self.table_dic:
            return

        del self.table_dic[oldcn]
        self.table_dic[cn] = lanId
 
        for i in range(self.id_start_row, self.max_row + 1):
            id_value = self.get_id(i)
            if id_value == lanId:
                self.sheet_src["{}{}".format('B', i)].value = cn
                break

    def save(self):
        # backup first
        if not os.path.exists(self.backup_folder):
            os.mkdir(self.backup_folder)
            shutil.copy(self.path_src, os.path.join(
                self.backup_folder, str(self.backup_order)))
            self.backup_order = (self.backup_order + 1) % self.max_backup_count

        self.workbook_src.save(self.path_src)

    def get_lan(self, row, lanType):
        col = get_column_letter(lanType+2)  # cn 2 en 3 zh 4 jp 5 ko 6
        value = self.sheet_src["{}{}".format(col, row)].value
        return str(value) if value is not None else ""

    def get_id(self, row):
        return int(self.sheet_src["{}{}".format('A', row)].value)
    pass


pass