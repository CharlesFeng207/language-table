import pymysql
import datetime
import Language_pb2
from Language_pb2 import LanguageTable, LanguageInfo
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter


def process_sql(func):
    con = pymysql.connect("localhost", "root", "", "language")
    cursor = con.cursor()
    try:
        func(con, cursor)
    except Exception as e:
        print(e)
    finally:
        cursor.close()
        con.close()


def save_history(info):
    with open("history.txt", "a+") as f:
        t = datetime.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
        f.write(f"{t} {info}</br>\n")


def query_statistic_info():
    count = 0

    def do(con, cursor):
        nonlocal count
        cursor.execute("select count(*) from language;")
        count = cursor.fetchone()[0]

    process_sql(do)

    return f"文本数量: {count}"


def query_id(cn):
    result = 0

    def do(con, cursor):
        nonlocal result
        rows = cursor.execute("select lanId from language where lan0=%s", cn)
        if rows > 0:
            result = cursor.fetchone()[0]

    process_sql(do)
    return result


def query_cn(lanId):
    result = None

    def do(con, cursor):
        nonlocal result
        rows = cursor.execute(
            "select lan0 from language where lanId=%s", lanId)
        if rows > 0:
            result = cursor.fetchone()[0]

    process_sql(do)
    return result

def query_all(lanId):
    result = None

    def do(con, cursor):
        nonlocal result
        rows = cursor.execute(
            "select * from language where lanId=%s", lanId)
        if rows > 0:
            result = cursor.fetchone()

    process_sql(do)
    return result


def insert_cn(cn):
    result = 0

    def do(con, cursor):
        nonlocal result
        rows = cursor.execute("INSERT INTO language (lan0) VALUES ( %s)", cn)
        con.commit()
        if rows > 0:
            result = query_id(cn)

    process_sql(do)

    if result != 0:
        save_history(f"录入 \"{cn}\"，id为{result}")

    return result


def edit_txt(lanId, lanType, newTxt):
    result = False

    col = "lan" + str(lanType)

    def do(con, cursor):
        nonlocal result
        rows = cursor.execute(
            f"UPDATE language SET {col}=%s WHERE lanId=%s;", (newTxt, lanId))
        con.commit()
        result = rows > 0

    process_sql(do)

    if result:
        save_history(f"编辑 {lanId} -> {newTxt}, lan:{lanType}")

    return result


def delete_lanId(lanId):
    result = False

    def do(con, cursor):
        nonlocal result
        rows = cursor.execute("DELETE FROM language WHERE lanId=%s;", lanId)
        con.commit()
        result = rows > 0

    process_sql(do)

    if result:
        save_history(f"删除Id {lanId}")

    return result


def export_proto_table():
    result = None

    def do(con, cursor):
        nonlocal result
        cursor.execute("SELECT * FROM language;")
        result = cursor.fetchall()

    process_sql(do)

    proto_table = LanguageTable()

    # id:0 "" "" ""
    zeroInfo = LanguageInfo()
    zeroInfo.id = 0
    for lantype in Language_pb2.LanguageType.values():
        zeroInfo.content.append("".encode('utf-8'))
    proto_table.infos.append(zeroInfo)

    for data in result:
        info = LanguageInfo()
        info.id = data[0]
        for lantype in Language_pb2.LanguageType.values():
            # 日文韩文暂时未能准备好，强制转换为英文
            if lantype == Language_pb2.ko or lantype == Language_pb2.jp:
                lantype = Language_pb2.en
            
            t = data[lantype+1] if data[lantype+1] is not None else ""
            info.content.append(t.encode('utf-8'))
        proto_table.infos.append(info)
        pass

    return proto_table


def export_workbook():
    result = None

    def do(con, cursor):
        nonlocal result
        cursor.execute("SELECT * FROM language;")
        result = cursor.fetchall()

    process_sql(do)
    nwb = Workbook()
    nsheet = nwb.active

    for i, data in enumerate(result):
        nsheet[f"A{i+1}"].value = data[0]  # id
        for lantype in Language_pb2.LanguageType.values():
            col = get_column_letter(lantype+2)
            nsheet[f"{col}{i+1}"].value = data[lantype+1]
            nsheet.column_dimensions[col].width = 50
    return nwb


if __name__ == "__main__":
    # export_workbook()
    pass
