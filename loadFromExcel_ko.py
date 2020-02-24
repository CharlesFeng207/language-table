# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, colors
import requests
from requests import Response
import os
import sys
import pymysql
import Language_pb2
import model


if __name__ == "__main__":

    path_src = "Language3.xlsx"
    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    start_row = 1
    lantype = Language_pb2.ko
    pending = []
    for row in range(start_row, sheet_src.max_row + 1):
        v = sheet_src[f"A{row}"].value
        if v is None:
            continue
        
        lanId = int(sheet_src[f"A{row}"].value)
        cell_ko = sheet_src[f"F{row}"]

        newTxt = str(cell_ko.value) if cell_ko.value is not None else ""
        p = (lanId, "", newTxt, lantype)
        pending.append(p)
    
    print(pending[0:10])
    print(len(pending))

    if input("press y to continue...") == "y":
        for lanId, oldTxt, newTxt, lantype in pending:
            print(lanId, oldTxt, newTxt)
            model.edit_txt(lanId, lantype, newTxt)
        pass
        

       


    
