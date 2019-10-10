# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, colors
import requests
from requests import Response
import os
import sys
import pymysql
import Language_pb2
import model


if __name__ == "__main__":
    db = pymysql.connect("localhost", "root", "", "language")
    cursor = db.cursor()
    cursor.execute("SELECT * FROM language;")
    t = cursor.fetchall()
    result = {i[0]: i for i in t}

    cursor.close()
    db.close()

    path_src = "Language2.xlsx"
    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    start_row = 1
    pending = []
    for row in range(start_row, sheet_src.max_row + 1):
        v = sheet_src[f"A{row}"].value
        if v is None:
            continue
        
        lanId = int(sheet_src[f"A{row}"].value)

        cell_cn = sheet_src[f"B{row}"]
        if cell_cn.fill.end_color.index == colors.BLACK: # only colored cell
            continue

        if lanId in result:
            for lantype in Language_pb2.LanguageType.values():
                cell = sheet_src[f"{get_column_letter(lantype + 2)}{row}"]
                # if cell.fill.end_color.index == colors.BLACK: # only colored cell
                #     continue

                if lantype is not Language_pb2.en and lantype is not Language_pb2.ch:
                    continue

                newTxt = str(cell.value) if cell.value is not None else ""

                oldTxt = result[lanId][lantype+1]

                if oldTxt != newTxt:
                    p = (lanId, oldTxt, newTxt, lantype)
                    print(p)
                    pending.append(p)

    print(len(pending))

    if input("press y to continue...") == "y":
        for lanId, oldTxt, newTxt, lantype in pending:
            print(lanId, oldTxt, newTxt)
            model.edit_txt(lanId, lantype, newTxt)
        pass
        

       


    
